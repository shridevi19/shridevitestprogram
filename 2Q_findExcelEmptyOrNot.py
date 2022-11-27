import os
import openpyxl
import os.path
import shutil
from datetime import *
import datetime
import excelFileCopyToAnotherPath

folder_location = r'\\cw01.contiwan.com\Root\Loc\lndp\didk7905\DOORS_Metrics\TEMP\Usha\Test_EmptyExcel_Files'
# file_list = create_file_list(folder_location)
folder = "LogFile"
folderPath = os.path.join(folder_location, folder)
# file_exists = os.path.exists(folderPath + "\log.txt")
# if(file_exists):
# os.remove(folderPath + "\log.txt")
# else:
folder_exists = os.path.exists(folderPath)
print("folder_exists--", folder_exists, folderPath)
if (folder_exists):
    print("if--")
    shutil.rmtree(folderPath, ignore_errors=False, onerror=None)
    os.makedirs(folderPath)
else:
    print("else--")
    os.makedirs(folderPath)

today_date = datetime.date.today()
new_today_date = today_date.strftime("%d-%m-%Y")
d1, m1, y1 = [int(x) for x in new_today_date.split('-')]
b1 = date(y1, m1, d1)
print("b1--", b1)


def readfile(fullpath):
    f = open(folderPath + "\log.txt", 'a+')
    f1 = open(folderPath + "\logDate.txt", 'a+')
    matchedDesc = ""
    notmatchedDesc = ""
    try:
        firstRwempty = False
        wb = openpyxl.load_workbook(fullpath)
        ws = wb[wb.sheetnames[0]]
        for r in range(1, 3):
            for c in range(1, 2):
                if r == 1 and c == 1:
                    s = ws.cell(r, c).value
                    if s is None:
                        firstRwempty = True
                        print("Header Blank--", fullpath)
                        f.write("Blank, %s \n" % fullpath)
                        # wb.close()
                    else:
                        pass

                if r == 2 and c == 1:
                    s = ws.cell(r, c).value
                    if s is None:
                        if not firstRwempty:
                            print("2nd Row Blank--", fullpath)
                            f.write("Blank, %s \n" % fullpath)
                        # wb.close()
                    else:
                        pass

        ws = wb.worksheets[0].values
        header = next(ws)  # get the header row
        my_data = []
        index = 0
        for row in ws:
            my_data.append(dict(zip(header, row)))
        for data in my_data:
            index += 1
            val = data['Exported on']
            val1 = val.strftime("%d-%m-%Y")
            d2, m2, y2 = [int(x) for x in val1.split('-')]
            b2 = date(y2, m2, d2)
            # print("b2--",b2)
            # print(b2, b1)
            if (b2 == b1):
                matchedDesc = "Matched"
                # print("MAtched")
            else:
                notmatchedDesc = "Not Matched"
                # print("Not MAtched")

        # ival=len(my_data) - 1
        index1 = index - 1
        # print(len(my_data) - 1, index1)
        if (index1 == len(my_data) - 1):
            # print("The last element of list using loop : ",notmatchedDesc)
            if (notmatchedDesc == "Not Matched"):
                f1.write("Todays date not available, '%s','%s' \n" % (fullpath, b1))
    except OSError as error:
        print("Directory '%s' can not be created %s" % error)
    except ValueError:
        print("File Corrupted--", fullpath)
        f.write("File Corrupted, %s \n" % fullpath)
    finally:
        f.close()
        f1.close()
        notmatchedDesc = ""
        matchedDesc = ""


for root, dirs, files in os.walk(folder_location):
    # print("dirs--",dirs)
    for name in files:
        if name.endswith((".xlsx", ".xls")):
            # print("root--", root, name)
            fullpath = root + "\\" + name
            # print("fullpath", fullpath)
            readfile(fullpath)


excelFileCopyToAnotherPath.deleteFun()
